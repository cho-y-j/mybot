"""
data.go.kr 전국동시지방선거 fileData XLSX → election_results_dong 임포트.

다운로드 (예시 — 8회 2022):
    curl -L -o /tmp/nec_2022_local.xlsx \\
      "https://www.data.go.kr/cmm/cmm/fileDownload.do?atchFileId=FILE_000000003157459&fileDetailSn=1&insertDataPrcus=N"

실행:
    cd backend
    python -m scripts.import_dong_xlsx --file /tmp/nec_2022_local.xlsx --year 2022 --number 8 --sido 충청북도

XLSX 구조 (sheet '구·시·군의장' 예시):
  Row1: 시도명/구시군명/선거구(구시군)/읍면동명/구분/선거인수/투표수/후보자별 득표수
  Row2: (None) ... 후보1/후보2/...
  Row3: 선거구 첫 행 — 후보명+정당 ('더불어민주당\\n송재봉', '국민의힘\\n이범석', ...)
  Row4: 합계
  Row5: 거소투표
  Row6: 관외사전투표
  Row7+: <읍면동> + '소계'
  Row8+: <읍면동> + '관내사전투표'
  Row9+: <읍면동> + '선거일투표'
  ... 다음 읍면동 ...
"""
import argparse
import asyncio
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import openpyxl

from app.database import async_session_factory
from app.elections.history_models import ElectionResultDong
from sqlalchemy import select, delete

# 시트 → election_type 매핑
SHEET_TO_TYPE = {
    "시·도지사": "governor",
    "구·시·군의장": "mayor",
    "시·도의회의원": "metro_council",
    "구·시·군의회의원": "basic_council",
    "교육감": "superintendent",
}


def _parse_party_name(cell: str) -> tuple[str, str]:
    """`'더불어민주당\\n송재봉'` → ('더불어민주당', '송재봉')."""
    if not cell or not isinstance(cell, str):
        return "", ""
    s = cell.strip()
    if "\n" in s:
        party, name = s.split("\n", 1)
        return party.strip(), name.strip()
    return "", s


def _to_int(v) -> int:
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).replace(",", "").strip()
    if not s or s == "-":
        return 0
    try:
        return int(s)
    except ValueError:
        return 0


def parse_sheet(ws, sido_filter: str | None = None) -> list[dict]:
    """시트 한 장을 읽어 (sigungu, eup_myeon_dong, candidate, party, ...) 리스트 반환.

    헤더 처리:
      - 한 선거구(시·군·구) 시작 → '시도/구시군/(선거구)/'' / '' / ''/ ''/ '정당\\n후보명' …'
      - 그 다음 행: 합계 / 거소투표 / 관외사전투표 / 읍면동별 …
    """
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # 첫 헤더 row에서 컬럼 위치 추정
    header = rows[0]
    # 컬럼명 검색
    def find_col(name):
        for i, c in enumerate(header):
            if c and str(c).strip() == name:
                return i
        return -1

    col_sido = find_col("시도명")
    # 구·시·군의장 시트엔 두 컬럼: '구시군명'(청주시) + '선거구(구시군)'(청주시상당구).
    # 후자가 더 정확한 단위(자치구 또는 통합시 일반구)이므로 우선 사용.
    col_sgg = find_col("선거구(구시군)")
    if col_sgg == -1:
        col_sgg = find_col("구시군명")
    if col_sgg == -1:
        col_sgg = find_col("선거구명")  # 시·도지사 sheet
    col_emd = find_col("읍면동명")
    col_gb = find_col("구분")

    if col_sido == -1 or col_emd == -1:
        # fallback by position (시도지사 sheet)
        col_sido = 0
        col_sgg = 1
        col_emd = 2
        col_gb = 3

    out = []
    current_candidates: list[tuple[str, str]] = []  # [(party, name)]
    current_sigungu: str | None = None

    # 데이터는 row 3부터 시작 (row 0=header, row 1=서브헤더 None, row 2=빈줄)
    for r in rows[2:]:
        if not r:
            continue

        sido = r[col_sido] if col_sido < len(r) else None
        sgg = r[col_sgg] if col_sgg < len(r) else None
        emd = r[col_emd] if col_emd < len(r) else None
        gb = r[col_gb] if col_gb < len(r) else None

        # 후보 헤더 row: emd/gb가 비어있고 후보 컬럼에 정당+후보명
        # 후보 컬럼은 '후보자별 득표수' 다음 (선거인수/투표수 2칸 다음부터)
        # 첫 번째 데이터 컬럼: col_gb + 3 (구분 + 선거인수 + 투표수)
        cand_start = col_gb + 3 if col_gb >= 0 else 6

        # 후보 헤더 row 판정: emd/gb 모두 비고 후보 컬럼에 텍스트
        if (emd is None or emd == "") and (gb is None or gb == ""):
            sample = r[cand_start] if cand_start < len(r) else None
            if sample and isinstance(sample, str) and sample.strip():
                # 헤더 row 추출 — 정당+후보명("정당\n후보명") 또는 후보명만(교육감)
                current_sigungu = sgg
                current_candidates = []
                for i in range(cand_start, len(r)):
                    cell = r[i]
                    if not cell or (isinstance(cell, str) and cell.strip() in ("", "\n")):
                        # 끝 도달 — 빈 셀 만나면 멈춤
                        # 단 교육감처럼 후보 6명 후 빈 컬럼은 정상이므로 그냥 break
                        break
                    party, name = _parse_party_name(str(cell))
                    if name:
                        current_candidates.append((party, name))
                continue

        # 데이터 row: emd가 있고 gb가 '소계' (또는 합계인 경우 별도)
        if not current_candidates or not sido:
            continue

        # sido 필터
        if sido_filter and str(sido).strip() != sido_filter:
            continue

        # 합계/거소/관외사전 행 무시 — 우리는 동별만
        if not emd or str(emd).strip() in ("", "합계"):
            continue

        # gb='소계' 인 경우만 (사전+선거일+거소 통합 = 동 전체)
        if gb is None or str(gb).strip() != "소계":
            continue

        # 후보별 득표 추출
        for idx, (party, name) in enumerate(current_candidates):
            col = cand_start + idx
            if col >= len(r):
                break
            votes = _to_int(r[col])
            out.append({
                "sido": str(sido).strip(),
                "sigungu": str(sgg).strip() if sgg else None,
                "eup_myeon_dong": str(emd).strip(),
                "candidate_name": name,
                "party": party or "무소속",
                "votes": votes,
            })

    return out


async def import_xlsx(file: str, year: int, number: int, sido: str | None):
    wb = openpyxl.load_workbook(file, read_only=True)
    print(f"=== {file} ({year}회) {sido or '전국'} ===")

    async with async_session_factory() as db:
        for sheet_name in wb.sheetnames:
            etype = SHEET_TO_TYPE.get(sheet_name)
            if not etype:
                continue
            print(f"\n--- {sheet_name} ({etype}) ---")
            ws = wb[sheet_name]
            records = parse_sheet(ws, sido_filter=sido)
            print(f"  parsed: {len(records)} (sido_filter={sido})")
            if not records:
                continue

            # 기존 데이터 삭제 후 재import (멱등)
            del_q = delete(ElectionResultDong).where(
                ElectionResultDong.election_year == year,
                ElectionResultDong.election_type == etype,
            )
            if sido:
                del_q = del_q.where(ElectionResultDong.region_sido == sido)
            await db.execute(del_q)

            # vote_rate 계산: 동별 후보 합계로 정규화
            from collections import defaultdict
            grouped = defaultdict(list)
            for rec in records:
                key = (rec["sigungu"], rec["eup_myeon_dong"])
                grouped[key].append(rec)

            inserted = 0
            for (sgg, emd), recs in grouped.items():
                total = sum(r["votes"] for r in recs) or 1
                max_votes = max(r["votes"] for r in recs) if recs else 0
                for rec in recs:
                    db.add(ElectionResultDong(
                        tenant_id=None,
                        election_number=number,
                        election_type=etype,
                        election_year=year,
                        region_sido=rec["sido"],
                        region_sigungu=sgg,
                        eup_myeon_dong=emd,
                        candidate_name=rec["candidate_name"],
                        party=rec["party"],
                        votes=rec["votes"],
                        vote_rate=round(rec["votes"] / total * 100, 2),
                        is_winner=(rec["votes"] == max_votes and max_votes > 0),
                    ))
                    inserted += 1
            await db.commit()
            print(f"  inserted: {inserted}")


async def main():
    p = argparse.ArgumentParser()
    p.add_argument("--file", required=True)
    p.add_argument("--year", type=int, required=True)
    p.add_argument("--number", type=int, required=True, help="회차 (예: 8)")
    p.add_argument("--sido", default=None, help="시도 필터 (예: 충청북도)")
    args = p.parse_args()
    await import_xlsx(args.file, args.year, args.number, args.sido)


if __name__ == "__main__":
    asyncio.run(main())
